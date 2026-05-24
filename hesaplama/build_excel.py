"""
Türkiye Limited Şirket - 2026 Yılı Vergi Hesaplama Tablosu Üreteci
Bu script formüllü XLSX dosyasını oluşturur.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

OUT = "/home/user/yasarhasimoglu.github.io/hesaplama/Limited_Sirket_Vergi_Hesaplama_2026.xlsx"

# ---- Stil yardımcıları ----
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12)
SUBHEAD_FILL = PatternFill("solid", fgColor="2E75B6")
SUBHEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
RESULT_FILL = PatternFill("solid", fgColor="E2EFDA")
TOTAL_FILL = PatternFill("solid", fgColor="C6E0B4")
WARN_FILL = PatternFill("solid", fgColor="FCE4D6")
BOLD = Font(bold=True)
ITALIC = Font(italic=True, color="595959")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

FMT_TL = '#,##0.00" ₺"'
FMT_PCT = '0.00%'
FMT_INT = '#,##0'


def set_header(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=cols[0], end_row=row, end_column=cols[1])
    c = ws.cell(row=row, column=cols[0], value=text)
    c.fill = HEADER_FILL
    c.font = HEADER_FONT
    c.alignment = CENTER
    ws.row_dimensions[row].height = 24


def set_subhead(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=cols[0], end_row=row, end_column=cols[1])
    c = ws.cell(row=row, column=cols[0], value=text)
    c.fill = SUBHEAD_FILL
    c.font = SUBHEAD_FONT
    c.alignment = CENTER
    ws.row_dimensions[row].height = 20


def label(ws, row, col, text, bold=True):
    c = ws.cell(row=row, column=col, value=text)
    if bold:
        c.font = BOLD
    c.alignment = LEFT
    c.border = BORDER
    return c


def input_cell(ws, row, col, value, fmt=FMT_TL):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = INPUT_FILL
    c.number_format = fmt
    c.border = BORDER
    c.alignment = RIGHT
    return c


def formula_cell(ws, row, col, formula, fmt=FMT_TL, fill=None, bold=False):
    c = ws.cell(row=row, column=col, value=formula)
    c.number_format = fmt
    if fill:
        c.fill = fill
    if bold:
        c.font = BOLD
    c.border = BORDER
    c.alignment = RIGHT
    return c


def note(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = ITALIC
    c.alignment = LEFT


wb = Workbook()

# =====================================================================
# SAYFA 1 - GİRDİLER (Parametreler)
# =====================================================================
ws = wb.active
ws.title = "1. Girdiler"
ws.sheet_view.showGridLines = False

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 55
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 60

set_header(ws, 1, (2, 5), "TÜRKİYE LİMİTED ŞİRKETİ – 2026 VERGİ HESAPLAMA TABLOSU")
ws.cell(row=2, column=2, value="Sarı hücreler = girdi (düzenleyebilirsiniz). Yeşil hücreler = otomatik hesaplanır.").font = ITALIC

# ---- 1.A Gelir & Maliyet ----
r = 4
set_subhead(ws, r, (2, 5), "1.A  Yıllık Gelir, Maliyet ve KDV Oranları"); r += 1
ws.cell(row=r, column=2, value="Parametre").font = BOLD
ws.cell(row=r, column=3, value="Değer").font = BOLD
ws.cell(row=r, column=4, value="Birim").font = BOLD
ws.cell(row=r, column=5, value="Açıklama").font = BOLD
for col in range(2, 6):
    ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="D9E1F2")
    ws.cell(row=r, column=col).alignment = CENTER
r += 1

# Satış (KDV hariç) - kullanıcının 1 milyon faturası varsayımı
label(ws, r, 2, "Yıllık satış / hasılat (KDV HARİÇ)")
input_cell(ws, r, 3, 1000000); ws.cell(row=r, column=4, value="₺").alignment = CENTER
note(ws, r, 5, "Kullanıcı varsayımı: 1.000.000 ₺ fatura kesimi (KDV hariç tutar).")
SALES_ROW = r; r += 1

label(ws, r, 2, "Satış KDV oranı")
input_cell(ws, r, 3, 0.20, FMT_PCT); ws.cell(row=r, column=4, value="%").alignment = CENTER
note(ws, r, 5, "Standart oran 2026 itibarıyla %20. Gıda %10, kitap %1 olabilir.")
SALES_VAT_RATE_ROW = r; r += 1

label(ws, r, 2, "Yıllık alış / gider (KDV HARİÇ, indirilebilir)")
input_cell(ws, r, 3, 350000); ws.cell(row=r, column=4, value="₺").alignment = CENTER
note(ws, r, 5, "Mal/hizmet alımı, kira, fatura vb. KDV'si indirilebilen giderler.")
COST_ROW = r; r += 1

label(ws, r, 2, "Alış KDV oranı (ortalama)")
input_cell(ws, r, 3, 0.20, FMT_PCT); ws.cell(row=r, column=4, value="%").alignment = CENTER
note(ws, r, 5, "Alımların ağırlıklı ortalama KDV oranı.")
COST_VAT_RATE_ROW = r; r += 1

label(ws, r, 2, "KDV'siz / belgesiz giderler (yıllık)")
input_cell(ws, r, 3, 0); ws.cell(row=r, column=4, value="₺").alignment = CENTER
note(ws, r, 5, "KKEG dışı ama KDV içermeyen giderler (örn. SGK primleri zaten ayrı).")
NONVAT_COST_ROW = r; r += 1

label(ws, r, 2, "Kanunen kabul edilmeyen giderler (KKEG)")
input_cell(ws, r, 3, 0); ws.cell(row=r, column=4, value="₺").alignment = CENTER
note(ws, r, 5, "Cezalar, özel binek araç giderlerinin bir kısmı vb. → Kurumlar Vergisi matrahına eklenir.")
KKEG_ROW = r; r += 1

# ---- 1.B Kira ----
r += 1
set_subhead(ws, r, (2, 5), "1.B  İşyeri Kirası ve Stopaj"); r += 1

label(ws, r, 2, "Aylık brüt işyeri kirası (gerçek kişiden)")
input_cell(ws, r, 3, 0); ws.cell(row=r, column=4, value="₺/ay").alignment = CENTER
note(ws, r, 5, "Gerçek kişi mal sahibinden kiralanıyorsa kiracı stopaj keser.")
RENT_MONTHLY_ROW = r; r += 1

label(ws, r, 2, "İşyeri kira stopaj oranı (GVK 94/5)")
input_cell(ws, r, 3, 0.20, FMT_PCT); ws.cell(row=r, column=4, value="%").alignment = CENTER
note(ws, r, 5, "2026 itibarıyla %20 (geçici düzenleme sona ererse).")
RENT_WHT_ROW = r; r += 1

# ---- 1.C Çalışanlar ----
r += 1
set_subhead(ws, r, (2, 5), "1.C  Çalışan / Personel Maliyeti"); r += 1

label(ws, r, 2, "Çalışan sayısı")
input_cell(ws, r, 3, 0, FMT_INT); ws.cell(row=r, column=4, value="kişi").alignment = CENTER
note(ws, r, 5, "0 girilirse personel maliyeti hesaplanmaz. Çalışan yoksa SGK primi yoktur.")
EMP_COUNT_ROW = r; r += 1

label(ws, r, 2, "Ortalama aylık BRÜT maaş (kişi başı)")
input_cell(ws, r, 3, 30000); ws.cell(row=r, column=4, value="₺/ay").alignment = CENTER
note(ws, r, 5, "Brüt maaş üzerinden tüm kesintiler ve işveren maliyeti hesaplanır.")
GROSS_SALARY_ROW = r; r += 1

label(ws, r, 2, "SGK işçi payı")
input_cell(ws, r, 3, 0.14, FMT_PCT); ws.cell(row=r, column=4, value="%").alignment = CENTER
note(ws, r, 5, "Standart %14 (5510 SK).")
SGK_EMP_ROW = r; r += 1

label(ws, r, 2, "İşsizlik sigortası işçi payı")
input_cell(ws, r, 3, 0.01, FMT_PCT); ws.cell(row=r, column=4, value="%").alignment = CENTER
ISSIZLIK_EMP_ROW = r; r += 1

label(ws, r, 2, "SGK işveren payı (5 puan indirim DAHİL %15,5)")
input_cell(ws, r, 3, 0.155, FMT_PCT); ws.cell(row=r, column=4, value="%").alignment = CENTER
note(ws, r, 5, "Normalde %20,5; 5510/Ek-81 indiriminden yararlanan işverenler %15,5 öder.")
SGK_ER_ROW = r; r += 1

label(ws, r, 2, "İşsizlik sigortası işveren payı")
input_cell(ws, r, 3, 0.02, FMT_PCT); ws.cell(row=r, column=4, value="%").alignment = CENTER
ISSIZLIK_ER_ROW = r; r += 1

label(ws, r, 2, "Damga vergisi oranı (ücret)")
input_cell(ws, r, 3, 0.00759, FMT_PCT); ws.cell(row=r, column=4, value="%").alignment = CENTER
note(ws, r, 5, "Binde 7,59 (488 SK).")
DAMGA_ROW = r; r += 1

label(ws, r, 2, "Aylık asgari ücret (brüt) – 2026 varsayımı")
input_cell(ws, r, 3, 30000); ws.cell(row=r, column=4, value="₺/ay").alignment = CENTER
note(ws, r, 5, "Asgari ücretten gelir/damga istisnası için kullanılır. 2026 rakamı netleştiğinde güncelleyin.")
MIN_WAGE_ROW = r; r += 1

# ---- 1.D Kurumlar Vergisi ----
r += 1
set_subhead(ws, r, (2, 5), "1.D  Kurumlar Vergisi ve Kâr Dağıtımı"); r += 1

label(ws, r, 2, "Kurumlar vergisi oranı")
input_cell(ws, r, 3, 0.25, FMT_PCT); ws.cell(row=r, column=4, value="%").alignment = CENTER
note(ws, r, 5, "2026 için genel oran %25. Bankalar/finans %30; halka açık (BIST %20+ fiili dolaşım) %23.")
CIT_RATE_ROW = r; r += 1

label(ws, r, 2, "Geçici (peşin) vergi oranı")
input_cell(ws, r, 3, 0.25, FMT_PCT); ws.cell(row=r, column=4, value="%").alignment = CENTER
note(ws, r, 5, "3 aylık dönemler hâlinde ödenir, yıl sonunda kurumlar vergisinden mahsup edilir.")
ADV_CIT_ROW = r; r += 1

label(ws, r, 2, "Kâr dağıtımı yapılacak mı?")
c = input_cell(ws, r, 3, "EVET", FMT_TL); c.number_format = 'General'; c.alignment = CENTER
note(ws, r, 5, "EVET / HAYIR. EVET ise kalan net kâr üzerinden temettü stopajı hesaplanır.")
DIV_FLAG_ROW = r; r += 1

label(ws, r, 2, "Temettü (kâr payı) stopaj oranı")
input_cell(ws, r, 3, 0.15, FMT_PCT); ws.cell(row=r, column=4, value="%").alignment = CENTER
note(ws, r, 5, "2025 sonu itibarıyla %15 (önceden %10 idi).")
DIV_WHT_ROW = r; r += 1

# ---- 1.E Diğer ----
r += 1
set_subhead(ws, r, (2, 5), "1.E  Diğer / Muafiyetler"); r += 1

label(ws, r, 2, "Asgari kurumlar vergisi oranı (yurt içi)")
input_cell(ws, r, 3, 0.10, FMT_PCT); ws.cell(row=r, column=4, value="%").alignment = CENTER
note(ws, r, 5, "7524 SK ile getirilen yurt içi asgari kurumlar vergisi (indirim öncesi matraha %10).")
MIN_CIT_ROW = r; r += 1

label(ws, r, 2, "Yeni kurulan şirket istisnası (ilk 3 yıl)?")
c = input_cell(ws, r, 3, "HAYIR", FMT_TL); c.number_format = 'General'; c.alignment = CENTER
note(ws, r, 5, "Asgari kurumlar vergisi yeni kurulanlar için ilk 3 yıl uygulanmaz (EVET = muaf).")
NEW_FIRM_ROW = r; r += 1

# Define named ranges to make formulas readable
named = {
    "Satis": f"'1. Girdiler'!$C${SALES_ROW}",
    "SatisKDV": f"'1. Girdiler'!$C${SALES_VAT_RATE_ROW}",
    "Alis": f"'1. Girdiler'!$C${COST_ROW}",
    "AlisKDV": f"'1. Girdiler'!$C${COST_VAT_RATE_ROW}",
    "KDVsizGider": f"'1. Girdiler'!$C${NONVAT_COST_ROW}",
    "KKEG": f"'1. Girdiler'!$C${KKEG_ROW}",
    "KiraAylik": f"'1. Girdiler'!$C${RENT_MONTHLY_ROW}",
    "KiraStopaj": f"'1. Girdiler'!$C${RENT_WHT_ROW}",
    "Calisan": f"'1. Girdiler'!$C${EMP_COUNT_ROW}",
    "BrutMaas": f"'1. Girdiler'!$C${GROSS_SALARY_ROW}",
    "SGKisci": f"'1. Girdiler'!$C${SGK_EMP_ROW}",
    "IssizlikIsci": f"'1. Girdiler'!$C${ISSIZLIK_EMP_ROW}",
    "SGKisveren": f"'1. Girdiler'!$C${SGK_ER_ROW}",
    "IssizlikIsveren": f"'1. Girdiler'!$C${ISSIZLIK_ER_ROW}",
    "Damga": f"'1. Girdiler'!$C${DAMGA_ROW}",
    "AsgariUcret": f"'1. Girdiler'!$C${MIN_WAGE_ROW}",
    "KVoran": f"'1. Girdiler'!$C${CIT_RATE_ROW}",
    "GeciciVergi": f"'1. Girdiler'!$C${ADV_CIT_ROW}",
    "TemettuVarMi": f"'1. Girdiler'!$C${DIV_FLAG_ROW}",
    "TemettuStopaj": f"'1. Girdiler'!$C${DIV_WHT_ROW}",
    "AsgariKVoran": f"'1. Girdiler'!$C${MIN_CIT_ROW}",
    "YeniSirket": f"'1. Girdiler'!$C${NEW_FIRM_ROW}",
}
for name, ref in named.items():
    wb.defined_names[name] = DefinedName(name=name, attr_text=ref)

# =====================================================================
# SAYFA 2 - KDV HESABI
# =====================================================================
ws2 = wb.create_sheet("2. KDV")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions['A'].width = 5
ws2.column_dimensions['B'].width = 55
ws2.column_dimensions['C'].width = 22
ws2.column_dimensions['D'].width = 55

set_header(ws2, 1, (2, 4), "KDV (KATMA DEĞER VERGİSİ) HESABI – YILLIK")
ws2.cell(row=2, column=2, value="KDV her ay beyan edilir. Burada yıllık toplam gösterilmiştir.").font = ITALIC

r = 4
set_subhead(ws2, r, (2, 4), "Hesaplanan KDV (Satışlardan)"); r += 1

label(ws2, r, 2, "Satış tutarı (KDV hariç)")
formula_cell(ws2, r, 3, "=Satis", fill=RESULT_FILL); note(ws2, r, 4, "Girdiler sayfasından")
r += 1
label(ws2, r, 2, "Satış KDV oranı")
formula_cell(ws2, r, 3, "=SatisKDV", FMT_PCT, fill=RESULT_FILL)
r += 1
label(ws2, r, 2, "Hesaplanan KDV (output VAT)")
formula_cell(ws2, r, 3, "=Satis*SatisKDV", fill=TOTAL_FILL, bold=True)
HESAP_KDV = r
note(ws2, r, 4, "Müşteriden tahsil edilen KDV")
r += 2

set_subhead(ws2, r, (2, 4), "İndirilecek KDV (Alışlardan)"); r += 1
label(ws2, r, 2, "Alış tutarı (KDV hariç)")
formula_cell(ws2, r, 3, "=Alis", fill=RESULT_FILL)
r += 1
label(ws2, r, 2, "Alış KDV oranı")
formula_cell(ws2, r, 3, "=AlisKDV", FMT_PCT, fill=RESULT_FILL)
r += 1
label(ws2, r, 2, "İndirilecek KDV (input VAT)")
formula_cell(ws2, r, 3, "=Alis*AlisKDV", fill=TOTAL_FILL, bold=True)
INDIRILECEK_KDV = r
note(ws2, r, 4, "Tedarikçilere ödenen ve mahsup edilen KDV")
r += 2

set_subhead(ws2, r, (2, 4), "Ödenecek KDV"); r += 1
label(ws2, r, 2, "Net Ödenecek KDV (yıllık)")
formula_cell(ws2, r, 3, f"=MAX(0, C{HESAP_KDV}-C{INDIRILECEK_KDV})", fill=TOTAL_FILL, bold=True)
ODENECEK_KDV = r
note(ws2, r, 4, "= Hesaplanan KDV - İndirilecek KDV. Negatifse devrolur, ödenmez.")
r += 1
label(ws2, r, 2, "Devreden KDV (varsa)")
formula_cell(ws2, r, 3, f"=MAX(0, C{INDIRILECEK_KDV}-C{HESAP_KDV})", fill=WARN_FILL, bold=True)
note(ws2, r, 4, "İndirilecek KDV > Hesaplanan KDV ise sonraki aya devreder.")
r += 1
label(ws2, r, 2, "Aylık ortalama ödenecek KDV")
formula_cell(ws2, r, 3, f"=C{ODENECEK_KDV}/12", fill=RESULT_FILL)
r += 2

note(ws2, r, 2, "ÖNEMLİ NOTLAR:"); ws2.cell(row=r, column=2).font = BOLD; r += 1
note(ws2, r, 2, "• KDV beyannamesi her ayın 28'ine kadar verilir, ödeme aynı tarihtir.")
r += 1
note(ws2, r, 2, "• 2026 standart KDV oranı %20'dir. İndirimli oranlar: gıda %10, kitap-eğitim %1, sağlık %10.")
r += 1
note(ws2, r, 2, "• Limited Şirket KDV mükellefidir; muafiyet yoktur (basit usul gerçek kişiler içindir).")
r += 1
note(ws2, r, 2, "• Tevkifatlı KDV (örn. yapım, danışmanlık) ayrıca dikkate alınmalıdır.")

# =====================================================================
# SAYFA 3 - ÇALIŞAN MALİYETİ
# =====================================================================
ws3 = wb.create_sheet("3. Çalışan Maliyeti")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions['A'].width = 5
ws3.column_dimensions['B'].width = 55
ws3.column_dimensions['C'].width = 22
ws3.column_dimensions['D'].width = 22
ws3.column_dimensions['E'].width = 55

set_header(ws3, 1, (2, 5), "ÇALIŞAN MALİYETİ – AYLIK VE YILLIK")
ws3.cell(row=2, column=2, value="Çalışan sayısı 0 ise tüm sonuçlar 0 olur. Gelir vergisi 2026 dilimleri ile hesaplanır.").font = ITALIC

r = 4
ws3.cell(row=r, column=3, value="KİŞİ BAŞI / AY").font = BOLD; ws3.cell(row=r, column=3).alignment = CENTER
ws3.cell(row=r, column=4, value="TOPLAM / YIL").font = BOLD; ws3.cell(row=r, column=4).alignment = CENTER
for col in (3, 4):
    ws3.cell(row=r, column=col).fill = PatternFill("solid", fgColor="D9E1F2")
r += 1

set_subhead(ws3, r, (2, 5), "A) İşçi Brüt Maaş ve Kesintiler"); r += 1

label(ws3, r, 2, "Brüt maaş")
formula_cell(ws3, r, 3, "=BrutMaas", fill=RESULT_FILL)
formula_cell(ws3, r, 4, f"=C{r}*Calisan*12", fill=RESULT_FILL)
BRUT_ROW = r; r += 1

label(ws3, r, 2, "(-) SGK işçi payı (%14)")
formula_cell(ws3, r, 3, f"=C{BRUT_ROW}*SGKisci", fill=RESULT_FILL)
formula_cell(ws3, r, 4, f"=C{r}*Calisan*12", fill=RESULT_FILL)
SGK_E_ROW = r; r += 1

label(ws3, r, 2, "(-) İşsizlik işçi payı (%1)")
formula_cell(ws3, r, 3, f"=C{BRUT_ROW}*IssizlikIsci", fill=RESULT_FILL)
formula_cell(ws3, r, 4, f"=C{r}*Calisan*12", fill=RESULT_FILL)
ISS_E_ROW = r; r += 1

label(ws3, r, 2, "Gelir vergisi matrahı (aylık)")
formula_cell(ws3, r, 3, f"=C{BRUT_ROW}-C{SGK_E_ROW}-C{ISS_E_ROW}", fill=RESULT_FILL)
formula_cell(ws3, r, 4, f"=C{r}*Calisan*12", fill=RESULT_FILL)
GV_MATRAH = r; r += 1

# Gelir vergisi - 2026 dilim varsayımı (2025'in yaklaşık %30 artırılmış hâli)
# 2025 dilimleri: 158k / 330k / 1.2M / 4.3M – 2026 için tahmini: 200k / 420k / 1.5M / 5.5M
label(ws3, r, 2, "(-) Gelir vergisi (yıllık kümülatif matraha göre)")
# Yıllık kümülatif matrah üzerinden hesaplama:
# y = yıllık matrah
# Dilimler 2026 tahmini: <200k: %15; 200-420k: %20+30000; 420-1.5M: %27+74000; 1.5-5.5M: %35+365600; >5.5M: %40+1.765.600
# Bunu tek formülde yapalım, yıllık olarak.
yil_matrah = f"C{GV_MATRAH}*12"
gv_formula = (
    f"=IF({yil_matrah}<=200000, {yil_matrah}*0.15, "
    f"IF({yil_matrah}<=420000, 30000+({yil_matrah}-200000)*0.20, "
    f"IF({yil_matrah}<=1500000, 74000+({yil_matrah}-420000)*0.27, "
    f"IF({yil_matrah}<=5500000, 365600+({yil_matrah}-1500000)*0.35, "
    f"1765600+({yil_matrah}-5500000)*0.40))))/12"
)
formula_cell(ws3, r, 3, gv_formula, fill=RESULT_FILL)
formula_cell(ws3, r, 4, f"=C{r}*Calisan*12", fill=RESULT_FILL)
GV_ROW = r
note(ws3, r, 5, "2026 dilim tahminleri: 200k/420k/1.5M/5.5M ₺ (resmî dilimler açıklandığında güncelleyin).")
r += 1

label(ws3, r, 2, "(-) Damga vergisi (binde 7,59)")
formula_cell(ws3, r, 3, f"=C{BRUT_ROW}*Damga", fill=RESULT_FILL)
formula_cell(ws3, r, 4, f"=C{r}*Calisan*12", fill=RESULT_FILL)
DAMGA_R = r; r += 1

# Asgari ücret istisnası: asgari ücretin GV+DV'si işverenden tahsil edilmez (çalışana iade edilir)
label(ws3, r, 2, "(+) Asgari ücret istisnası (GV+DV)")
# Asgari ücretten hesaplanan GV+DV → istisna
# Asgari ücret aylık brüt, SGK işçi+işsizlik düşülüp dilime giren kısım %15 GV + damga
agi_formula = (
    f"=IF(Calisan=0,0, "
    f"((AsgariUcret-AsgariUcret*(SGKisci+IssizlikIsci))*0.15 + AsgariUcret*Damga))"
)
formula_cell(ws3, r, 3, agi_formula, fill=RESULT_FILL)
formula_cell(ws3, r, 4, f"=C{r}*Calisan*12", fill=RESULT_FILL)
AGI_ROW = r
note(ws3, r, 5, "Asgari ücret üzerinden hesaplanan GV+DV istisna; net maaşa eklenir.")
r += 1

label(ws3, r, 2, "NET MAAŞ (çalışana ödenen)")
formula_cell(ws3, r, 3,
             f"=C{BRUT_ROW}-C{SGK_E_ROW}-C{ISS_E_ROW}-C{GV_ROW}-C{DAMGA_R}+C{AGI_ROW}",
             fill=TOTAL_FILL, bold=True)
formula_cell(ws3, r, 4, f"=C{r}*Calisan*12", fill=TOTAL_FILL, bold=True)
NET_ROW = r; r += 2

set_subhead(ws3, r, (2, 5), "B) İşveren Maliyetleri (Brüt Üzerinden)"); r += 1

label(ws3, r, 2, "SGK işveren payı (%15,5 indirimli / %20,5 normal)")
formula_cell(ws3, r, 3, f"=C{BRUT_ROW}*SGKisveren", fill=RESULT_FILL)
formula_cell(ws3, r, 4, f"=C{r}*Calisan*12", fill=RESULT_FILL)
SGK_IV = r; r += 1

label(ws3, r, 2, "İşsizlik işveren payı (%2)")
formula_cell(ws3, r, 3, f"=C{BRUT_ROW}*IssizlikIsveren", fill=RESULT_FILL)
formula_cell(ws3, r, 4, f"=C{r}*Calisan*12", fill=RESULT_FILL)
ISS_IV = r; r += 1

label(ws3, r, 2, "TOPLAM İŞVEREN MALİYETİ (kişi başı/ay – yıllık)")
formula_cell(ws3, r, 3, f"=C{BRUT_ROW}+C{SGK_IV}+C{ISS_IV}", fill=TOTAL_FILL, bold=True)
formula_cell(ws3, r, 4, f"=C{r}*Calisan*12", fill=TOTAL_FILL, bold=True)
TOPLAM_ISV = r
note(ws3, r, 5, "Şirket için tek bir çalışanın gerçek yıllık maliyeti.")
r += 2

set_subhead(ws3, r, (2, 5), "C) Muhtasar Beyanname ile Ödenecekler (Yıllık)"); r += 1
label(ws3, r, 2, "Çalışan gelir vergisi stopajı (yıllık, istisna sonrası)")
formula_cell(ws3, r, 3, f"=MAX(0, (D{GV_ROW}+D{DAMGA_R}) - D{AGI_ROW})", fill=RESULT_FILL)
note(ws3, r, 5, "Muhtasar beyanname ile her ay/3 ayda bir ödenir.")
MUHTASAR_UCRET = r; r += 1

# Kira stopajı
label(ws3, r, 2, "Kira stopajı (gerçek kişiden kira ödeniyorsa)")
formula_cell(ws3, r, 3, "=KiraAylik*12*KiraStopaj", fill=RESULT_FILL)
note(ws3, r, 5, "Aylık brüt kira × 12 × %20.")
MUHTASAR_KIRA = r; r += 1

label(ws3, r, 2, "TOPLAM MUHTASAR")
formula_cell(ws3, r, 3, f"=C{MUHTASAR_UCRET}+C{MUHTASAR_KIRA}", fill=TOTAL_FILL, bold=True)
MUHTASAR_TOPLAM = r; r += 2

note(ws3, r, 2, "MUAFİYET/AVANTAJ NOTLARI:"); ws3.cell(row=r, column=2).font = BOLD; r += 1
note(ws3, r, 2, "• Çalışan yoksa (Çalışan=0): SGK ve gelir vergisi stopajı yok, ancak ortaklar BAĞ-KUR (4/b) ödemekle yükümlüdür.")
r += 1
note(ws3, r, 2, "• Asgari ücretten hesaplanan gelir vergisi + damga vergisi istisnadır (tüm çalışanlar için).")
r += 1
note(ws3, r, 2, "• İlk defa işe alınan kadın/genç/engelli için ek SGK teşvikleri (4447 SK 19, 7103 SK) olabilir.")
r += 1
note(ws3, r, 2, "• 10'dan az çalışanı olanlar muhtasarı 3 ayda bir verebilir; aksi hâlde aylık.")

# =====================================================================
# SAYFA 4 - KURUMLAR VERGİSİ
# =====================================================================
ws4 = wb.create_sheet("4. Kurumlar Vergisi")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions['A'].width = 5
ws4.column_dimensions['B'].width = 55
ws4.column_dimensions['C'].width = 22
ws4.column_dimensions['D'].width = 60

set_header(ws4, 1, (2, 4), "KURUMLAR VERGİSİ HESABI – YILLIK")
ws4.cell(row=2, column=2, value="Ticari kâr → KKEG/istisnalarla mali kâra çevrilir → vergi hesaplanır.").font = ITALIC

r = 4
set_subhead(ws4, r, (2, 4), "A) Ticari Kâr Hesabı (Kazanç)"); r += 1

label(ws4, r, 2, "Satış geliri (KDV hariç)")
formula_cell(ws4, r, 3, "=Satis", fill=RESULT_FILL)
GELIR = r; r += 1

label(ws4, r, 2, "(-) Mal/hizmet alımları & giderler (KDV hariç)")
formula_cell(ws4, r, 3, "=Alis", fill=RESULT_FILL)
ALIS_G = r; r += 1

label(ws4, r, 2, "(-) KDV'siz/belgesiz indirilebilir giderler")
formula_cell(ws4, r, 3, "=KDVsizGider", fill=RESULT_FILL)
KDVSIZ_G = r; r += 1

label(ws4, r, 2, "(-) Yıllık kira gideri (brüt)")
formula_cell(ws4, r, 3, "=KiraAylik*12", fill=RESULT_FILL)
KIRA_G = r
note(ws4, r, 4, "Brüt kira gider, stopaj ayrıca düşülmez; sadece muhasebede gider.")
r += 1

label(ws4, r, 2, "(-) Toplam personel gideri (işveren maliyeti)")
formula_cell(ws4, r, 3, f"='3. Çalışan Maliyeti'!D{TOPLAM_ISV}", fill=RESULT_FILL)
PERS_G = r; r += 1

label(ws4, r, 2, "TİCARİ KÂR")
formula_cell(ws4, r, 3, f"=C{GELIR}-C{ALIS_G}-C{KDVSIZ_G}-C{KIRA_G}-C{PERS_G}", fill=TOTAL_FILL, bold=True)
TIC_KAR = r; r += 2

set_subhead(ws4, r, (2, 4), "B) Mali Kâr (Vergi Matrahı)"); r += 1
label(ws4, r, 2, "Ticari kâr")
formula_cell(ws4, r, 3, f"=C{TIC_KAR}", fill=RESULT_FILL)
r += 1
label(ws4, r, 2, "(+) Kanunen kabul edilmeyen giderler (KKEG)")
formula_cell(ws4, r, 3, "=KKEG", fill=RESULT_FILL)
KKEG_R = r; r += 1
label(ws4, r, 2, "(-) İstisna ve indirimler")
formula_cell(ws4, r, 3, 0, fill=INPUT_FILL)
IST_R = r
note(ws4, r, 4, "Manuel: AR-GE indirimi, iştirak kazancı istisnası, KOBİ birleşme istisnası vb.")
r += 1
label(ws4, r, 2, "VERGİ MATRAHI (Mali Kâr)")
formula_cell(ws4, r, 3, f"=MAX(0, C{TIC_KAR}+C{KKEG_R}-C{IST_R})", fill=TOTAL_FILL, bold=True)
MATRAH = r; r += 2

set_subhead(ws4, r, (2, 4), "C) Kurumlar Vergisi"); r += 1
label(ws4, r, 2, "Hesaplanan kurumlar vergisi (matrah × %25)")
formula_cell(ws4, r, 3, f"=C{MATRAH}*KVoran", fill=RESULT_FILL)
HESAP_KV = r; r += 1

label(ws4, r, 2, "Yurt içi asgari kurumlar vergisi (matrah × %10)")
formula_cell(ws4, r, 3,
             f'=IF(UPPER(TRIM(YeniSirket))="EVET",0, C{MATRAH}*AsgariKVoran)',
             fill=RESULT_FILL)
ASG_KV = r
note(ws4, r, 4, "7524 SK. Yeni kurulan şirketler ilk 3 yıl muaf. Hesaplanan KV bundan azsa fark ödenir.")
r += 1

label(ws4, r, 2, "ÖDENECEK KURUMLAR VERGİSİ (en büyük olanı)")
formula_cell(ws4, r, 3, f"=MAX(C{HESAP_KV}, C{ASG_KV})", fill=TOTAL_FILL, bold=True)
ODENECEK_KV = r; r += 1

label(ws4, r, 2, "Geçici (peşin) vergi (3 ayda bir, yıl içinde ödenen)")
formula_cell(ws4, r, 3, f"=C{ODENECEK_KV}", fill=RESULT_FILL)
note(ws4, r, 4, "3 aylık dönemlerde ödenir, yıl sonunda mahsup edilir; ek bir vergi DEĞİL.")
r += 2

set_subhead(ws4, r, (2, 4), "D) Net Kâr ve Temettü"); r += 1

label(ws4, r, 2, "VERGİ SONRASI NET KÂR (ortaklara dağıtılabilir)")
formula_cell(ws4, r, 3, f"=C{TIC_KAR}-C{ODENECEK_KV}", fill=TOTAL_FILL, bold=True)
NET_KAR = r; r += 1

label(ws4, r, 2, "Temettü stopajı (kâr dağıtımı yapılırsa)")
formula_cell(ws4, r, 3,
             f'=IF(UPPER(TRIM(TemettuVarMi))="EVET", MAX(0,C{NET_KAR})*TemettuStopaj, 0)',
             fill=RESULT_FILL)
TEMETTU_S = r
note(ws4, r, 4, "Net kâr ortaklara dağıtılırsa şirket %15 stopaj keser ve muhtasar ile öder.")
r += 1

label(ws4, r, 2, "ORTAKLARIN ELİNE GEÇEN NET TEMETTÜ")
formula_cell(ws4, r, 3,
             f'=IF(UPPER(TRIM(TemettuVarMi))="EVET", MAX(0,C{NET_KAR})-C{TEMETTU_S}, 0)',
             fill=TOTAL_FILL, bold=True)
ORTAK_NET = r; r += 1

note(ws4, r, 2, "Not: Gerçek kişi ortak temettüyü beyan eder; %15 stopaj mahsup edilir, dilime göre ek vergi çıkabilir.")
ws4.cell(row=r, column=2).font = ITALIC

# =====================================================================
# SAYFA 5 - STOPAJ ÖZETİ
# =====================================================================
ws5 = wb.create_sheet("5. Stopajlar")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions['A'].width = 5
ws5.column_dimensions['B'].width = 50
ws5.column_dimensions['C'].width = 22
ws5.column_dimensions['D'].width = 60

set_header(ws5, 1, (2, 4), "STOPAJ (KAYNAĞINDA KESİLEN VERGİ) ÖZETİ – YILLIK")

r = 3
set_subhead(ws5, r, (2, 4), "Muhtasar Beyanname Kalemleri"); r += 1

label(ws5, r, 2, "Ücret stopajı (çalışanlardan)")
formula_cell(ws5, r, 3, f"='3. Çalışan Maliyeti'!C{MUHTASAR_UCRET}", fill=RESULT_FILL)
UC_S = r; r += 1

label(ws5, r, 2, "Kira stopajı (işyeri kirası)")
formula_cell(ws5, r, 3, f"='3. Çalışan Maliyeti'!C{MUHTASAR_KIRA}", fill=RESULT_FILL)
KR_S = r
note(ws5, r, 4, "Aylık brüt kira × 12 × %20")
r += 1

label(ws5, r, 2, "Serbest meslek stopajı (varsa)")
formula_cell(ws5, r, 3, 0, fill=INPUT_FILL)
SM_S = r
note(ws5, r, 4, "SMM, avukat, mali müşavir gibi serbest meslek erbabına yapılan ödemelerde %20.")
r += 1

label(ws5, r, 2, "Temettü stopajı (kâr dağıtımı)")
formula_cell(ws5, r, 3, f"='4. Kurumlar Vergisi'!C{TEMETTU_S}", fill=RESULT_FILL)
TE_S = r; r += 1

label(ws5, r, 2, "TOPLAM MUHTASAR / YIL")
formula_cell(ws5, r, 3, f"=SUM(C{UC_S}:C{TE_S})", fill=TOTAL_FILL, bold=True)
TOPLAM_M = r; r += 2

note(ws5, r, 2, "Muhtasar beyanname genelde takip eden ayın 26'sına kadar verilir, 26'sına kadar ödenir.")
ws5.cell(row=r, column=2).font = ITALIC

# =====================================================================
# SAYFA 6 - ÖZET
# =====================================================================
ws6 = wb.create_sheet("6. ÖZET")
ws6.sheet_view.showGridLines = False
ws6.column_dimensions['A'].width = 5
ws6.column_dimensions['B'].width = 55
ws6.column_dimensions['C'].width = 25
ws6.column_dimensions['D'].width = 18
ws6.column_dimensions['E'].width = 50

set_header(ws6, 1, (2, 5), "ŞİRKETİN YILLIK TOPLAM VERGİ YÜKÜ – ÖZET TABLO")
ws6.cell(row=2, column=2, value="Aşağıdaki tüm rakamlar Girdiler sayfasındaki varsayımlardan otomatik hesaplanır.").font = ITALIC

r = 4
ws6.cell(row=r, column=2, value="Vergi / Kalem").font = BOLD
ws6.cell(row=r, column=3, value="Yıllık Tutar").font = BOLD
ws6.cell(row=r, column=4, value="Hasılat %").font = BOLD
ws6.cell(row=r, column=5, value="Beyan / Ödeme Dönemi").font = BOLD
for col in range(2, 6):
    ws6.cell(row=r, column=col).fill = PatternFill("solid", fgColor="D9E1F2")
    ws6.cell(row=r, column=col).alignment = CENTER
    ws6.cell(row=r, column=col).border = BORDER
r += 1

# Satırlar
items = [
    ("Ödenecek KDV (yıllık net)", f"='2. KDV'!C{ODENECEK_KDV}", "Aylık, 28'i"),
    ("Kurumlar vergisi", f"='4. Kurumlar Vergisi'!C{ODENECEK_KV}", "Yıllık (Nisan); geçici 3 ayda bir"),
    ("Ücret stopajı (muhtasar)", f"='5. Stopajlar'!C{UC_S}", "Aylık/3 aylık, 26'sı"),
    ("Kira stopajı (muhtasar)", f"='5. Stopajlar'!C{KR_S}", "Aylık/3 aylık, 26'sı"),
    ("Serbest meslek stopajı", f"='5. Stopajlar'!C{SM_S}", "Aylık/3 aylık, 26'sı"),
    ("Temettü stopajı", f"='5. Stopajlar'!C{TE_S}", "Dağıtım yapıldığında"),
    ("SGK işveren primi (yıllık)", f"='3. Çalışan Maliyeti'!D{SGK_IV}", "Aylık, takip eden ayın sonu"),
    ("İşsizlik işveren primi (yıllık)", f"='3. Çalışan Maliyeti'!D{ISS_IV}", "Aylık, takip eden ayın sonu"),
]

first_data = r
for name, formula, period in items:
    label(ws6, r, 2, name, bold=False)
    formula_cell(ws6, r, 3, formula, fill=RESULT_FILL)
    formula_cell(ws6, r, 4, f"=C{r}/Satis", FMT_PCT, fill=RESULT_FILL)
    c = ws6.cell(row=r, column=5, value=period); c.alignment = LEFT; c.border = BORDER
    r += 1
last_data = r - 1

label(ws6, r, 2, "ŞİRKETİN ÖDEDİĞİ TOPLAM (vergi + SGK)")
formula_cell(ws6, r, 3, f"=SUM(C{first_data}:C{last_data})", fill=TOTAL_FILL, bold=True)
formula_cell(ws6, r, 4, f"=C{r}/Satis", FMT_PCT, fill=TOTAL_FILL, bold=True)
ws6.cell(row=r, column=5, value="").border = BORDER
TOPLAM_HEPSI = r
r += 2

# Mini panel
set_subhead(ws6, r, (2, 5), "Hızlı Panel"); r += 1
panels = [
    ("Yıllık satış (KDV hariç)", "=Satis"),
    ("Yıllık alış (KDV hariç)", "=Alis"),
    ("Ticari kâr (vergi öncesi)", f"='4. Kurumlar Vergisi'!C{TIC_KAR}"),
    ("Vergi matrahı (mali kâr)", f"='4. Kurumlar Vergisi'!C{MATRAH}"),
    ("Vergi sonrası net kâr", f"='4. Kurumlar Vergisi'!C{NET_KAR}"),
    ("Ortakların eline geçen (temettü sonrası)", f"='4. Kurumlar Vergisi'!C{ORTAK_NET}"),
    ("Toplam personel maliyeti (işveren)", f"='3. Çalışan Maliyeti'!D{TOPLAM_ISV}"),
    ("Çalışan sayısı", "=Calisan"),
]
for n, f in panels:
    label(ws6, r, 2, n, bold=False)
    fmt = FMT_INT if n == "Çalışan sayısı" else FMT_TL
    formula_cell(ws6, r, 3, f, fmt, fill=RESULT_FILL)
    r += 1

r += 1
note(ws6, r, 2, "→ İlk sayfadaki sarı hücreleri değiştirerek tüm tabloyu kendi senaryonuza uyarlayabilirsiniz.")
ws6.cell(row=r, column=2).font = BOLD

# =====================================================================
# SAYFA 7 - NOTLAR VE ORANLAR
# =====================================================================
ws7 = wb.create_sheet("7. Notlar & Oranlar")
ws7.sheet_view.showGridLines = False
ws7.column_dimensions['A'].width = 5
ws7.column_dimensions['B'].width = 40
ws7.column_dimensions['C'].width = 90

set_header(ws7, 1, (2, 3), "2026 VERGİ ORANLARI VE ÖNEMLİ NOTLAR")

r = 3
content = [
    ("KDV oranları", "Standart %20 | İndirimli %10 (gıda, sağlık vb.) | %1 (temel gıda, kitap, eğitim)"),
    ("Kurumlar vergisi", "Genel %25 | Bankalar/finansal kuruluşlar %30 | Halka açık şirketler (BIST'te min %20 fiili dolaşım) %23"),
    ("Yurt içi asgari kurumlar vergisi", "7524 SK ile matrahın %10'undan az olamaz. Yeni kurulanlar ilk 3 yıl muaf."),
    ("Geçici vergi", "Üçer aylık dönemler için %25; yıl sonunda kurumlar vergisinden mahsup edilir."),
    ("Temettü stopajı", "2025 sonu itibarıyla %15 (önceden %10 idi). Cumhurbaşkanı kararıyla değişebilir."),
    ("Gelir vergisi (ücret) – 2026 tahmini dilimler",
        "<200.000 ₺: %15 | 200-420 bin: %20 | 420 bin-1,5 M: %27 | 1,5-5,5 M: %35 | >5,5 M: %40 (resmî dilimler açıklandığında güncelleyin)"),
    ("Damga vergisi (ücret)", "Binde 7,59 — brüt maaş üzerinden."),
    ("SGK işçi", "%14 (emeklilik+sağlık) + %1 işsizlik = %15"),
    ("SGK işveren", "%20,5 normal; 5510 SK 5 puan indirimiyle %15,5 + %2 işsizlik = %17,5 (indirimli)"),
    ("Kira stopajı", "İşyeri kirası gerçek kişiden ise %20 (GVK 94/5). Tüzel kişiden kiralanıyorsa stopaj yok."),
    ("Serbest meslek stopajı", "Genel %20 (SMMM, avukat, doktor vb. ödemelerde)."),
    ("KDV mükellefiyeti", "Limited şirket otomatik KDV mükellefidir; basit usul / muafiyet uygulanmaz."),
    ("BAĞ-KUR (4/b)", "Limited şirket ortağı çalışan olmasa bile 4/b'ye tabidir; aylık prim öder. Bu tablo şirket vergilerini hesaplar, ortak bireysel BAĞ-KUR'u ayrıdır."),
    ("Defter tutma", "Bilanço esasına göre defter tutar (e-Defter zorunlu). e-Fatura, e-Arşiv ciroya göre zorunlu olabilir."),
    ("Beyan dönemleri",
        "KDV: aylık 28'i | Muhtasar: aylık veya 3 aylık 26'sı | Geçici vergi: 3 ayda bir 17'si | Kurumlar vergisi: Nisan ayı 30'a kadar | SGK: takip eden ayın son günü"),
    ("Asgari ücret istisnası", "Tüm çalışanlar için, brüt asgari ücretten hesaplanan GV ve damga vergisi istisnadır."),
    ("Yeni teşvik / istisnalar",
        "AR-GE, tasarım, ihracat (KVK 32/A), yatırım teşvik (indirimli KV), genç girişimci (GV içinden), KOBİ birleşme istisnası → uygulanıyorsa Sayfa 4 'istisna' alanına girilir."),
    ("Yasal Uyarı",
        "Bu tablo bilgilendirme amaçlıdır; nihai vergi yükümlülüğü için Mali Müşavir / YMM ile teyit ediniz. 2026 yılı resmi oranları yıl başında Gelir İdaresi Başkanlığı tarafından yayımlanır."),
]

for k, v in content:
    label(ws7, r, 2, k)
    c = ws7.cell(row=r, column=3, value=v)
    c.alignment = LEFT
    c.border = BORDER
    ws7.row_dimensions[r].height = max(20, 18 * (1 + len(v) // 90))
    r += 1

# Freeze panes ve sekme renkleri
ws.sheet_properties.tabColor = "FFC000"
ws2.sheet_properties.tabColor = "2E75B6"
ws3.sheet_properties.tabColor = "70AD47"
ws4.sheet_properties.tabColor = "C00000"
ws5.sheet_properties.tabColor = "7030A0"
ws6.sheet_properties.tabColor = "000000"
ws7.sheet_properties.tabColor = "808080"

for s in [ws, ws2, ws3, ws4, ws5, ws6, ws7]:
    s.freeze_panes = "A4"

wb.save(OUT)
print(f"OK -> {OUT}")
